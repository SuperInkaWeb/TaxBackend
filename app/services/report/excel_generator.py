import io
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.services.reconciliation.engine import ReconciliationOutput, IGV_DIFF_THRESHOLD, TIPO_LABELS


RED_FILL     = PatternFill("solid", fgColor="FFC7CE")
RED_FONT     = Font(color="9C0006")
AMBER_FILL   = PatternFill("solid", fgColor="FFEB9C")
AMBER_FONT   = Font(color="9C6500")
GREEN_FILL   = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT   = Font(color="006100")
DIFF_FILL    = PatternFill("solid", fgColor="FFD966")
HEADER_FILL  = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT  = Font(color="FFFFFF", bold=True, size=11)
TITLE_FILL   = PatternFill("solid", fgColor="1F4E78")
TITLE_FONT   = Font(color="FFFFFF", bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11, color="1F4E78")
BOLD         = Font(bold=True)
GRAY_FILL    = PatternFill("solid", fgColor="F2F2F2")
NOTE_FONT    = Font(italic=True, size=9, color="595959")

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"), right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"), bottom=Side(style="thin", color="BFBFBF"),
)

NUM_FMT = "#,##0.00"

EXCEL_FORMAT_LIMIT  = 50_000
EXCEL_B_LIMIT       = 10_000
EXCEL_D_LIMIT       = 10_000
EXCEL_MAX_DATA_ROWS = 1_048_575


def _set_header_row(ws, headers: list[str], row: int = 1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 28


def _finish_sheet(ws, n_cols: int, num_cols: list[int] | None = None, widths: dict[int, int] | None = None):
    """Autofiltro + panel congelado + formato numérico + anchos."""
    last_col = get_column_letter(n_cols)
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
    ws.freeze_panes = "A2"

    if num_cols:
        for col_idx in num_cols:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = NUM_FMT

    if widths:
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    else:
        for col in range(1, n_cols + 1):
            max_len = max(
                (len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row, 200) + 1)),
                default=8,
            )
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 40)


def _alert_style(cell, es_roja: bool):
    cell.fill = RED_FILL if es_roja else AMBER_FILL
    cell.font = RED_FONT if es_roja else AMBER_FONT


def generate_excel(
    output: ReconciliationOutput,
    empresa_nombre: str,
    ruc: str,
    periodo: str,
    tipo_libro: str,
    propuesta_generada: datetime | None = None,
) -> bytes:
    wb = Workbook()

    cnt_a = len(output.scenario_a)
    cnt_b = len(output.scenario_b)
    cnt_c = len(output.scenario_c)
    cnt_d = len(output.scenario_d)

    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:G3")
    title = ws["B2"]
    title.value = "REPORTE DE CONCILIACIÓN SIRE"
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws["B2:G3"]:
        for c in row:
            c.fill = TITLE_FILL

    r = 5
    ws.cell(row=r, column=2, value="DATOS GENERALES").font = SECTION_FONT
    r += 1
    datos = [
        ("Empresa",       empresa_nombre),
        ("RUC",           ruc),
        ("Periodo",       f"{periodo[:4]}/{periodo[4:]}"),
        ("Tipo de libro", "Ventas (RVIE)" if tipo_libro == "ventas" else "Compras (RCE)"),
    ]
    if propuesta_generada is not None:
        datos.append(("Propuesta SUNAT", f"generada el {propuesta_generada.astimezone().strftime('%d/%m/%Y %H:%M')}"))
    datos.append(("Generado", datetime.now().strftime("%d/%m/%Y %H:%M")))
    for label, value in datos:
        ws.cell(row=r, column=2, value=label).font = BOLD
        ws.cell(row=r, column=3, value=value)
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="RESULTADOS").font = SECTION_FONT
    r += 1
    resultados = [
        ("A — En tu archivo, no en SUNAT",          cnt_a, "Comprobantes que declaraste pero SUNAT no tiene registrados en todo el mes", False),
        ("B — En SUNAT, no en tu archivo",          cnt_b, "Comprobantes que SUNAT tiene en las fechas de tu archivo y tú no enviaste", False),
        ("C — En ambos, con diferencias",           cnt_c, "Comprobantes encontrados en ambos pero con fecha o montos distintos", False),
        ("D — Coinciden sin diferencias",           cnt_d, "Comprobantes validados: idénticos en tu archivo y en SUNAT", True),
    ]
    for label, count, desc, es_bueno in resultados:
        ws.cell(row=r, column=2, value=label).font = BOLD
        cell_cnt = ws.cell(row=r, column=3, value=count)
        cell_cnt.font = Font(bold=True, size=12)
        cell_cnt.alignment = Alignment(horizontal="center")
        if count > 0:
            cell_cnt.fill = GREEN_FILL if es_bueno else AMBER_FILL
        ws.cell(row=r, column=4, value=desc).font = NOTE_FONT
        r += 1

    total_comparados = cnt_a + cnt_c + cnt_d
    if total_comparados:
        ws.cell(row=r, column=2, value="Tasa de conciliación").font = BOLD
        cell_tasa = ws.cell(row=r, column=3, value=cnt_d / total_comparados)
        cell_tasa.number_format = "0.00%"
        cell_tasa.font = Font(bold=True, size=12, color="006100")
        cell_tasa.alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4, value="Porcentaje de tus comprobantes que coinciden exactamente con SUNAT").font = NOTE_FONT
        r += 1

    r += 1
    ws.cell(row=r, column=2, value="Total diferencia de IGV").font = BOLD
    cell_igv = ws.cell(row=r, column=3, value=output.igv_diferencia_total)
    cell_igv.number_format = f'"S/" {NUM_FMT}'
    cell_igv.font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=2, value="Alertas rojas").font = BOLD
    cell_al = ws.cell(row=r, column=3, value="SÍ" if output.tiene_alertas_rojas else "NO")
    cell_al.alignment = Alignment(horizontal="center")
    if output.tiene_alertas_rojas:
        _alert_style(cell_al, es_roja=True)
        cell_al.font = Font(bold=True, color="9C0006")
    r += 1

    r += 1
    ws.cell(row=r, column=2, value="LEYENDA").font = SECTION_FONT
    r += 1
    cell = ws.cell(row=r, column=2, value="ROJO")
    _alert_style(cell, es_roja=True)
    cell.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=3, value=f"Riesgo de IGV: falta o diferencia con impacto mayor a S/ {IGV_DIFF_THRESHOLD:.2f}. Revisar primero.")
    r += 1
    cell = ws.cell(row=r, column=2, value="ÁMBAR")
    _alert_style(cell, es_roja=False)
    cell.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=3, value="Revisar: inconsistencia sin impacto directo de IGV (fecha, montos sin IGV).")
    r += 1
    cell = ws.cell(row=r, column=2, value="Celda resaltada")
    cell.fill = DIFF_FILL
    cell.alignment = Alignment(horizontal="center")
    ws.cell(row=r, column=3, value="En la pestaña C: el valor de esa celda es el que difiere entre tu archivo y SUNAT.")
    r += 1

    notas = []
    if output.total_excluidos:
        detalle = ", ".join(
            f"{TIPO_LABELS.get(t, t)} ({t}): {c:,}"
            for t, c in sorted(output.excluidos_por_tipo.items(), key=lambda x: -x[1])
        )
        notas.append((
            f"Registros de tu archivo excluidos de la comparación: {output.total_excluidos:,}",
            f"SUNAT no incluye estos tipos en la propuesta → {detalle}",
        ))
    if output.csv_duplicados:
        notas.append((
            f"Filas duplicadas en tu archivo: {output.csv_duplicados:,}",
            "Misma clave Tipo+Serie+Número repetida; se usó la última aparición.",
        ))
    if output.sunat_duplicados:
        notas.append((
            f"Filas duplicadas en la propuesta SUNAT: {output.sunat_duplicados:,}",
            "Se usó la última aparición.",
        ))
    if cnt_b > EXCEL_B_LIMIT:
        notas.append((
            f"La pestaña B muestra {EXCEL_B_LIMIT:,} de {cnt_b:,} registros",
            "Ordenados por IGV descendente. El listado completo está en el CSV descargable.",
        ))
    if cnt_d > EXCEL_MAX_DATA_ROWS:
        notas.append((
            f"La pestaña D muestra {EXCEL_MAX_DATA_ROWS:,} de {cnt_d:,} registros",
            "Excel no admite más filas por hoja. El listado completo está en el CSV del Escenario D.",
        ))
    elif cnt_d > EXCEL_D_LIMIT:
        notas.append((
            f"La pestaña D contiene el detalle completo ({cnt_d:,} comprobantes OK)",
            "El mismo listado está disponible como CSV del Escenario D para análisis externo.",
        ))

    if notas:
        r += 1
        ws.cell(row=r, column=2, value="NOTAS").font = SECTION_FONT
        r += 1
        for nota, detalle in notas:
            ws.cell(row=r, column=2, value=nota).font = BOLD
            ws.cell(row=r, column=3, value=detalle).font = NOTE_FONT
            r += 1

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 70

    ws_a = wb.create_sheet("A - Solo en tu archivo")
    _set_header_row(ws_a, [
        "Tipo", "Serie", "Número", "Fecha emisión",
        "Base imponible", "IGV", "Importe total", "Estado (tu archivo)", "Alerta",
    ])
    use_fmt_a = cnt_a <= EXCEL_FORMAT_LIMIT
    for row_idx, rec in enumerate(output.scenario_a, 2):
        values = [
            rec.tipo_cdp, rec.serie, rec.numero, rec.fecha_emision,
            rec.base_imponible, rec.igv, rec.importe_total,
            rec.status_description,
            "ROJO" if rec.es_alerta_roja else "ÁMBAR",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_a.cell(row=row_idx, column=col_idx, value=val)
            if use_fmt_a:
                cell.border = THIN_BORDER
                if col_idx == 8 and "RECHAZ" in str(val).upper():
                    cell.font = Font(bold=True, color="9C0006")
                elif col_idx == 9:
                    _alert_style(cell, rec.es_alerta_roja)
                    cell.alignment = Alignment(horizontal="center")
    _finish_sheet(ws_a, 9, num_cols=[5, 6, 7])

    ws_b = wb.create_sheet("B - Solo en SUNAT")
    _set_header_row(ws_b, [
        "Tipo", "Serie", "Número", "Fecha emisión SUNAT",
        "Base imponible", "IGV", "Importe total", "Alerta",
    ])
    recs_b_sorted = sorted(output.scenario_b, key=lambda x: (not x.es_alerta_roja, -x.igv_sunat))
    recs_b = recs_b_sorted[:EXCEL_B_LIMIT]
    use_fmt_b = len(recs_b) <= EXCEL_FORMAT_LIMIT

    for row_idx, rec in enumerate(recs_b, 2):
        values = [
            rec.tipo_cdp, rec.serie, rec.numero, rec.fecha_emision,
            rec.base_imponible_sunat, rec.igv_sunat, rec.importe_total_sunat,
            "ROJO" if rec.es_alerta_roja else "ÁMBAR",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_b.cell(row=row_idx, column=col_idx, value=val)
            if use_fmt_b:
                cell.border = THIN_BORDER
                if col_idx == 8:
                    _alert_style(cell, rec.es_alerta_roja)
                    cell.alignment = Alignment(horizontal="center")
    _finish_sheet(ws_b, 8, num_cols=[5, 6, 7])

    ws_c = wb.create_sheet("C - Diferencias")
    _set_header_row(ws_c, [
        "Tipo", "Serie", "Número",
        "Fecha (tu archivo)", "Fecha (SUNAT)",
        "Base imp. (tu archivo)", "Base imp. (SUNAT)",
        "IGV (tu archivo)", "IGV (SUNAT)",
        "Total (tu archivo)", "Total (SUNAT)",
        "Exonerado (tu archivo)", "Exonerado (SUNAT)",
        "Inafecto (tu archivo)", "Inafecto (SUNAT)",
        "Campos con diferencia", "Alerta",
    ])
    use_fmt_c = cnt_c <= EXCEL_FORMAT_LIMIT

    CAMPO_LABELS = {
        "fecha": "Fecha", "base_imponible": "Base imponible",
        "igv": "IGV", "importe_total": "Importe total",
        "mto_exonerado": "Exonerado", "mto_inafecto": "Inafecto",
    }
    CAMPO_COLS = {
        "fecha":          (4, 5),
        "base_imponible": (6, 7),
        "igv":            (8, 9),
        "importe_total":  (10, 11),
        "mto_exonerado":  (12, 13),
        "mto_inafecto":   (14, 15),
    }
    CAMPO_ORDER = ["fecha", "base_imponible", "igv", "importe_total", "mto_exonerado", "mto_inafecto"]

    for row_idx, rec in enumerate(output.scenario_c, 2):
        campos = rec.campos_diferentes
        campos_txt = ", ".join(CAMPO_LABELS[c] for c in CAMPO_ORDER if c in campos)

        values = [
            rec.tipo_cdp, rec.serie, rec.numero,
            rec.fecha_emision_empresa, rec.fecha_emision_sunat,
            rec.base_imponible_empresa, rec.base_imponible_sunat,
            rec.igv_empresa, rec.igv_sunat,
            rec.importe_total_empresa, rec.importe_total_sunat,
            rec.mto_exonerado_empresa, rec.mto_exonerado_sunat,
            rec.mto_inafecto_empresa, rec.mto_inafecto_sunat,
            campos_txt,
            "ROJO" if rec.es_alerta_roja else "ÁMBAR",
        ]
        diff_cols = {col for campo in campos for col in CAMPO_COLS.get(campo, ())}

        for col_idx, val in enumerate(values, 1):
            cell = ws_c.cell(row=row_idx, column=col_idx, value=val)
            if use_fmt_c:
                cell.border = THIN_BORDER
                if col_idx in diff_cols:
                    cell.fill = DIFF_FILL
                    cell.font = BOLD
                elif col_idx == 17:
                    _alert_style(cell, rec.es_alerta_roja)
                    cell.alignment = Alignment(horizontal="center")
    _finish_sheet(ws_c, 17, num_cols=[6, 7, 8, 9, 10, 11, 12, 13, 14, 15])

    ws_d = wb.create_sheet("D - Coinciden OK")
    _set_header_row(ws_d, [
        "Tipo", "Serie", "Número",
        "Fecha (tu archivo)", "Fecha (SUNAT)",
        "Base imp. (tu archivo)", "Base imp. (SUNAT)",
        "IGV (tu archivo)", "IGV (SUNAT)",
        "Total (tu archivo)", "Total (SUNAT)",
        "Exonerado (tu archivo)", "Exonerado (SUNAT)",
        "Inafecto (tu archivo)", "Inafecto (SUNAT)",
    ])
    use_fmt_d = cnt_d <= EXCEL_FORMAT_LIMIT
    recs_d = output.scenario_d[:EXCEL_MAX_DATA_ROWS]
    for row_idx, rec in enumerate(recs_d, 2):
        values = [
            rec.tipo_cdp, rec.serie, rec.numero,
            rec.fecha_emision_empresa, rec.fecha_emision_sunat,
            rec.base_imponible_empresa, rec.base_imponible_sunat,
            rec.igv_empresa, rec.igv_sunat,
            rec.importe_total_empresa, rec.importe_total_sunat,
            rec.mto_exonerado_empresa, rec.mto_exonerado_sunat,
            rec.mto_inafecto_empresa, rec.mto_inafecto_sunat,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_d.cell(row=row_idx, column=col_idx, value=val)
            if use_fmt_d:
                cell.border = THIN_BORDER
    _finish_sheet(ws_d, 15, num_cols=[6, 7, 8, 9, 10, 11, 12, 13, 14, 15] if use_fmt_d else None)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_csv_b(output: ReconciliationOutput) -> bytes:
    """CSV completo del Escenario B — solo se genera cuando B supera EXCEL_B_LIMIT."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        "Tipo CDP", "Serie", "Numero", "Fecha Emision SUNAT",
        "Base Imponible", "IGV", "Importe Total", "Alerta",
    ])
    for rec in output.scenario_b:
        w.writerow([
            rec.tipo_cdp, rec.serie, rec.numero, rec.fecha_emision,
            rec.base_imponible_sunat, rec.igv_sunat, rec.importe_total_sunat,
            "ROJO" if rec.es_alerta_roja else "AMBAR",
        ])
    return buf.getvalue().encode("utf-8-sig")


def generate_csv_d(output: ReconciliationOutput) -> bytes:
    """CSV completo del Escenario D — solo se genera cuando D supera EXCEL_D_LIMIT."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        "Tipo CDP", "Serie", "Numero",
        "Fecha (Empresa)", "Fecha (SUNAT)",
        "Base Imponible (Empresa)", "Base Imponible (SUNAT)",
        "IGV (Empresa)", "IGV (SUNAT)",
        "Importe Total (Empresa)", "Importe Total (SUNAT)",
        "Exonerado (Empresa)", "Exonerado (SUNAT)",
        "Inafecto (Empresa)", "Inafecto (SUNAT)",
    ])
    for rec in output.scenario_d:
        w.writerow([
            rec.tipo_cdp, rec.serie, rec.numero,
            rec.fecha_emision_empresa, rec.fecha_emision_sunat,
            rec.base_imponible_empresa, rec.base_imponible_sunat,
            rec.igv_empresa, rec.igv_sunat,
            rec.importe_total_empresa, rec.importe_total_sunat,
            rec.mto_exonerado_empresa, rec.mto_exonerado_sunat,
            rec.mto_inafecto_empresa, rec.mto_inafecto_sunat,
        ])
    return buf.getvalue().encode("utf-8-sig")
